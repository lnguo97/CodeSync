import json
from pathlib import Path
from typing import Literal
from uuid import uuid4

import streamlit as st
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models import Template


if 'db_session' not in st.session_state:
    DATABASE_URL = "sqlite:///./test.db"
    engine = create_engine(DATABASE_URL, echo=True)
    Session = sessionmaker(bind=engine)
    st.session_state.db_session = Session()


class TemplateUI:
    def __init__(self, template: Template) -> None:
        self.template = template

    def form(self, mode: Literal['create', 'update']) -> bool:
        if st.session_state.get('submit_template'):
            result, err_msg = self.validate(mode)
            if result:
                return True
            st.error(f'Error: {err_msg}.')
        st.text_input(
            'Template Name',
            value=self.template.name,
            key=f'name_{self.template.id}'
        )
        st.file_uploader(
            'Template File',
            type='xlsx',
            key=f'file_{self.template.id}'
        )
        st.button('Submit', key=f'submit_template')

    def validate(self, mode: Literal['create', 'update']) -> tuple[bool, str | None]:
        name = st.session_state.get(f'name_{self.template.id}')
        if not name.strip():
            return False, 'Template name cannot be empty'
        session = st.session_state.db_session
        all_templates = session.query(Template).all()
        if name in [t.name for t in all_templates] and mode == 'create':
            return False, f'Template "{name}" already exists'
        file = st.session_state.get(f'file_{self.template.id}')
        if file is None:
            return False, 'Template file cannot be empty'
        file_path = Path(f'template/{uuid4().hex}.xlsx')
        with open(file_path, mode='wb') as f:
            f.write(file.getvalue())
        wb = load_workbook(file_path)
        meta_data = dict()
        for sheet in wb.sheetnames:
            meta_data[sheet] = {
                tb_name: ref
                for tb_name, ref in wb[sheet].tables.items()
            }
        self.template.name = name
        self.template.file_path = str(file_path)
        self.template.meta_data = json.dumps(meta_data)
        return True, None

    @staticmethod
    @st.dialog('Create New Template')
    def create() -> None:
        if 'new_template' not in st.session_state:
            st.session_state.new_template = Template(
                name='', file_path='', meta_data=''
            )
        template_ui = TemplateUI(st.session_state.new_template)
        submitted = template_ui.form(mode='create')
        if submitted:
            new_template = st.session_state.pop('new_template')
            st.session_state.db_session.add(new_template)
            st.session_state.db_session.commit()
            st.rerun()

    def retrieve(self) -> None:
        st.write(template.name, template.file_path)

    def update(self) -> None:
        pass


st.title('Report Template')
st.button(
    'New Template',
    icon=':material/add:',
    on_click=TemplateUI.create
)
session = st.session_state.db_session
templates = session.query(Template).all()
for template in templates:
    template_ui = TemplateUI(template)
    template_ui.retrieve()
